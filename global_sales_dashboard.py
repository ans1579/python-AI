# -*- coding: utf-8 -*-
import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import os
import numpy as np

# ✅ .env 파일 자동 로드 추가
from dotenv import load_dotenv
load_dotenv()  # .env 파일 로드

# [ 환경 변수 및 Gemini API 설정 ]
GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY")
FIVECORE_BASE_URL = os.environ.get("FIVECORE_BASE_URL", "http://localhost:8090/5core")

# ==================== 페이지 설정 ====================
st.set_page_config(
    page_title="글로벌 판매 분석",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# ==================== CSS 스타일 ====================
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@300;400;500;600;700&display=swap');
    
    * { font-family: 'Noto Sans KR', sans-serif; }
    
    .header-section {
        background: linear-gradient(135deg, #111827 0%, #1f2937 40%, #111827 100%);
        padding: 24px 28px;
        border-radius: 14px;
        margin-bottom: 16px;
        color: #f9fafb;
        box-shadow: 0 10px 30px rgba(15, 23, 42, 0.4);
        border: 1px solid #374151;
    }
    .header-title {
        font-size: 24px;
        font-weight: 700;
        margin: 0 0 4px 0;
    }
    .header-subtitle {
        font-size: 13px;
        margin: 0;
        color: #d1d5db;
    }

    .control-panel {
        margin-bottom: 20px;
        padding: 16px 18px;
        border-radius: 10px;
        background: linear-gradient(135deg, #1f2937 0%, #111827 100%);
        border: 1px solid #4b5563;
    }
    .control-block-title {
        font-size: 11px;
        font-weight: 600;
        letter-spacing: 0.06em;
        text-transform: uppercase;
        color: #9ca3af;
        margin-bottom: 8px;
        display: block;
    }

    .kpi-card {
        background: white;
        padding: 18px;
        border-radius: 10px;
        text-align: center;
        box-shadow: 0 4px 12px rgba(0,0,0,0.08);
        border: 1px solid #e5e7eb;
    }
    .kpi-label { 
        color: #6b7280; 
        font-size: 12px; 
        font-weight: 500; 
        margin-bottom: 6px; 
    }
    .kpi-value { 
        color: #1f2937; 
        font-size: 28px; 
        font-weight: 700; 
    }
    .kpi-unit { 
        color: #9ca3af; 
        font-size: 11px; 
        margin-top: 4px; 
    }

    .section-title {
        font-size: 16px;
        font-weight: 600;
        color: #111827;
        margin: 20px 0 14px 0;
        padding-bottom: 8px;
        border-bottom: 2px solid #e5e7eb;
    }

    .comparison-box {
        background: linear-gradient(135deg, #f0f9ff 0%, #e0f2fe 100%);
        padding: 14px 16px;
        border-radius: 10px;
        border-left: 4px solid #0284c7;
        margin-bottom: 16px;
    }

    /* 라디오/체크박스 스타일 */
    .stRadio > label, .stCheckbox > label {
        font-size: 13px;
        font-weight: 500;
        color: #e5e7eb !important;
        margin-bottom: 4px;
    }
    .stRadio > div[role="radiogroup"],
    .stRadio > div[data-testid="stRadio"] {
        gap: 8px !important;
    }
</style>
""", unsafe_allow_html=True)

# ==================== 헤더 섹션 ====================
st.markdown("""
<div class="header-section">
    <div class="header-title">현대차 글로벌 판매 실적 분석 대시보드</div>
    <p class="header-subtitle">
        공장별(국내/해외) 판매 비중, 월별 추이, 모델별 성과, 연도별 비교를 한 화면에서 분석
    </p>
</div>
""", unsafe_allow_html=True)

# ==================== 컨트롤 패널 ====================
st.markdown('<div class="control-panel">', unsafe_allow_html=True)

col_ctrl1, col_ctrl2, col_ctrl3 = st.columns([1.2, 1.8, 1.5], gap="medium")

with col_ctrl1:
    st.markdown('<span class="control-block-title">기준 연도 선택</span>', unsafe_allow_html=True)
    selected_year = st.radio(
        label="year_selection",
        options=[2024, 2025],
        index=1,
        format_func=lambda x: f"{x}년",
        horizontal=True,
        label_visibility="collapsed"
    )

with col_ctrl2:
    st.markdown('<span class="control-block-title">연도별 비교 옵션</span>', unsafe_allow_html=True)
    compare_mode = st.radio(
        label="compare_selection",
        options=["단일 연도 보기", "이전 연도와 비교"],
        index=0,
        horizontal=True,
        label_visibility="collapsed"
    )
    show_comparison = compare_mode == "이전 연도와 비교"

with col_ctrl3:
    st.markdown('<span class="control-block-title">추가 옵션</span>', unsafe_allow_html=True)
    layout_mode = st.checkbox("상세 데이터 테이블", value=True)

st.markdown('</div>', unsafe_allow_html=True)

# ==================== 데이터 로드 ====================
@st.cache_data(ttl=600)
def load_data(year: int) -> pd.DataFrame | None:
    file = f"{year}_Global_Sales_Clean.xlsx"
    if os.path.exists(file):
        return pd.read_excel(file, sheet_name="Global Sales")
    return None

df = load_data(selected_year)
if df is None:
    st.error(f"{selected_year}년 데이터를 찾을 수 없습니다!")
    st.stop()

# 월 컬럼
months = ['Jan.', 'Feb.', 'Mar.', 'Apr.', 'May.', 'Jun.',
          'Jul.', 'Aug.', 'Sep.', 'Oct.', 'Nov.', 'Dec.']
months_ko = ['1월', '2월', '3월', '4월', '5월', '6월',
             '7월', '8월', '9월', '10월', '11월', '12월']

# 공장별 분리
domestic = df[df["Plant"] == "국내"]
export = df[df["Plant"] == "수출"]

domestic_total = domestic[months].sum().sum()
export_total = export[months].sum().sum()
grand_total = domestic_total + export_total

# 비교 데이터
df_compare = None
if show_comparison:
    base_year = 2024 if selected_year == 2025 else 2025
    df_compare = load_data(base_year)

# ==================== 1. 연도별 비교 분석 (맨 위) ====================
if show_comparison and df_compare is not None:
    st.markdown('<h2 class="section-title">연도별 비교 분석</h2>', unsafe_allow_html=True)

    domestic_compare = df_compare[df_compare["Plant"] == "국내"]
    export_compare = df_compare[df_compare["Plant"] == "수출"]

    domestic_total_compare = domestic_compare[months].sum().sum()
    export_total_compare = export_compare[months].sum().sum()
    base_total = domestic_total_compare + export_total_compare
    base_year = 2024 if selected_year == 2025 else 2025

    col_kpi1, col_kpi2, col_kpi3, col_kpi4 = st.columns(4, gap="large")

    with col_kpi1:
        growth = (grand_total - base_total) / base_total * 100
        st.metric(
            label="총 판매량 변화율",
            value=f"{growth:+.1f}%",
            delta=f"{grand_total - base_total:+,.0f} 대",
        )

    with col_kpi2:
        g_dom = (domestic_total - domestic_total_compare) / domestic_total_compare * 100
        st.metric(
            label="국내 공장 판매 변화율",
            value=f"{g_dom:+.1f}%",
            delta=f"{domestic_total - domestic_total_compare:+,.0f} 대",
        )

    with col_kpi3:
        g_exp = (export_total - export_total_compare) / export_total_compare * 100
        st.metric(
            label="해외 공장 판매 변화율",
            value=f"{g_exp:+.1f}%",
            delta=f"{export_total - export_total_compare:+,.0f} 대",
        )

    with col_kpi4:
        st.metric(
            label="모델 수 변화",
            value=df["Model"].nunique(),
            delta=df["Model"].nunique() - df_compare["Model"].nunique(),
        )

    st.markdown("---")

    monthly_domestic = domestic[months].sum()
    monthly_export = export[months].sum()
    monthly_domestic_compare = domestic_compare[months].sum()
    monthly_export_compare = export_compare[months].sum()

    fig_comparison = go.Figure()
    fig_comparison.add_trace(
        go.Bar(
            x=months_ko,
            y=monthly_domestic.values,
            name=f"{selected_year}년 국내",
            marker_color="#1e40af",
            hovertemplate="<b>%{x}</b><br>판매량: %{y:,.0f}대<extra></extra>",
        )
    )
    fig_comparison.add_trace(
        go.Bar(
            x=months_ko,
            y=monthly_domestic_compare.values,
            name=f"{base_year}년 국내",
            marker_color="#93c5fd",
            hovertemplate="<b>%{x}</b><br>판매량: %{y:,.0f}대<extra></extra>",
        )
    )
    fig_comparison.add_trace(
        go.Bar(
            x=months_ko,
            y=monthly_export.values,
            name=f"{selected_year}년 해외",
            marker_color="#f97316",
            hovertemplate="<b>%{x}</b><br>판매량: %{y:,.0f}대<extra></extra>",
        )
    )
    fig_comparison.add_trace(
        go.Bar(
            x=months_ko,
            y=monthly_export_compare.values,
            name=f"{base_year}년 해외",
            marker_color="#fed7aa",
            hovertemplate="<b>%{x}</b><br>판매량: %{y:,.0f}대<extra></extra>",
        )
    )

    fig_comparison.update_layout(
        barmode="group",
        height=420,
        yaxis_title="판매대수(대)",
        yaxis_tickformat=",",
        hovermode="x unified",
    )
    st.plotly_chart(fig_comparison, use_container_width=True)

    st.markdown("---")

# ==================== 2. 핵심 지표 ====================
st.markdown('<h2 class="section-title">핵심 지표</h2>', unsafe_allow_html=True)

col_kpi_main1, col_kpi_main2, col_kpi_main3, col_kpi_main4 = st.columns(4, gap="large")

with col_kpi_main1:
    st.markdown(
        f"""
    <div class="kpi-card">
        <div class="kpi-label">총 판매량</div>
        <div class="kpi-value">{grand_total:,.0f}</div>
        <div class="kpi-unit">대</div>
    </div>
    """,
        unsafe_allow_html=True,
    )

with col_kpi_main2:
    st.markdown(
        f"""
    <div class="kpi-card">
        <div class="kpi-label">국내 공장 판매</div>
        <div class="kpi-value">{domestic_total:,.0f}</div>
        <div class="kpi-unit">{domestic_total/grand_total*100:.1f}%</div>
    </div>
    """,
        unsafe_allow_html=True,
    )

with col_kpi_main3:
    st.markdown(
        f"""
    <div class="kpi-card">
        <div class="kpi-label">해외 공장 판매</div>
        <div class="kpi-value">{export_total:,.0f}</div>
        <div class="kpi-unit">{export_total/grand_total*100:.1f}%</div>
    </div>
    """,
        unsafe_allow_html=True,
    )

with col_kpi_main4:
    st.markdown(
        f"""
    <div class="kpi-card">
        <div class="kpi-label">등록 모델 수</div>
        <div class="kpi-value">{df['Model'].nunique()}</div>
        <div class="kpi-unit">종</div>
    </div>
    """,
        unsafe_allow_html=True,
    )

st.markdown("---")

# ==================== 3. 공장별 연간 판매 비중 ====================
st.markdown('<h2 class="section-title">연간 총 판매량 - 공장별 비중</h2>', unsafe_allow_html=True)

col_pie1, col_pie2 = st.columns([2, 1], gap="large")

with col_pie1:
    fig_pie = go.Figure(
        data=[
            go.Pie(
                labels=["국내 공장", "해외 공장"],
                values=[domestic_total, export_total],
                hole=0.4,
                marker=dict(colors=["#1e40af", "#f97316"]),
                textinfo="label+percent+value",
                hovertemplate="<b>%{label}</b><br>판매량: %{value:,.0f}대<br>비중: %{percent}<extra></extra>",
            )
        ]
    )
    fig_pie.update_layout(height=380, showlegend=True)
    st.plotly_chart(fig_pie, use_container_width=True)

with col_pie2:
    st.markdown(
        f"""
    <div class="comparison-box">
        <h4 style="color:#0284c7; margin-top:0;">공장별 판매 요약</h4>
        <table style="width:100%; border-collapse: collapse;">
            <tr style="border-bottom: 1px solid #bae6fd;">
                <td style="padding: 8px; font-weight: 600;">국내</td>
                <td style="padding: 8px; text-align: right; color: #1e40af; font-weight: 700;">{domestic_total:,.0f}</td>
            </tr>
            <tr style="border-bottom: 1px solid #bae6fd;">
                <td style="padding: 8px; font-weight: 600;">해외</td>
                <td style="padding: 8px; text-align: right; color: #f97316; font-weight: 700;">{export_total:,.0f}</td>
            </tr>
            <tr style="background: #eff6ff; font-weight: 700;">
                <td style="padding: 8px;">합계</td>
                <td style="padding: 8px; text-align: right; color: #1f2937;">{grand_total:,.0f}</td>
            </tr>
        </table>
    </div>
    """,
        unsafe_allow_html=True,
    )

st.markdown("---")

# ==================== 4. 월별 판매 추이 비교 ====================
st.markdown('<h2 class="section-title">월별 판매 추이 비교</h2>', unsafe_allow_html=True)

monthly_domestic = domestic[months].sum()
monthly_export = export[months].sum()

fig_line = go.Figure()
fig_line.add_trace(
    go.Scatter(
        x=months_ko,
        y=monthly_domestic.values,
        name="국내 공장",
        mode="lines+markers",
        line=dict(color="#1e40af", width=3),
        marker=dict(size=8),
        hovertemplate="<b>국내 공장</b><br>%{x}: %{y:,.0f}대<extra></extra>",
    )
)
fig_line.add_trace(
    go.Scatter(
        x=months_ko,
        y=monthly_export.values,
        name="해외 공장",
        mode="lines+markers",
        line=dict(color="#f97316", width=3),
        marker=dict(size=8),
        hovertemplate="<b>해외 공장</b><br>%{x}: %{y:,.0f}대<extra></extra>",
    )
)

fig_line.update_layout(
    height=420,
    hovermode="x unified",
    xaxis_title="월",
    yaxis_title="판매대수 (대)",
    yaxis_tickformat=",",
    legend=dict(x=0.01, y=0.99, bgcolor="rgba(255,255,255,0.8)"),
)
st.plotly_chart(fig_line, use_container_width=True)

st.markdown("**월별 판매량 상세:**")
monthly_summary = pd.DataFrame(
    {
        "월": months_ko,
        "국내": monthly_domestic.values.astype(int),
        "해외": monthly_export.values.astype(int),
    }
)
monthly_summary["합계"] = monthly_summary["국내"] + monthly_summary["해외"]
st.dataframe(monthly_summary.set_index("월"), use_container_width=True)

st.markdown("---")

# ==================== 5. 주력 모델 기여도 ====================
st.markdown('<h2 class="section-title">주력 모델 기여도 (Top 10)</h2>', unsafe_allow_html=True)

col_top1, col_top2 = st.columns(2, gap="large")

with col_top1:
    st.markdown("**국내 공장 - Top 10 모델**")
    top_domestic = domestic.nlargest(10, "Total")[["Model", "Total"]].sort_values("Total")
    fig_bar_domestic = px.bar(
        top_domestic,
        x="Total",
        y="Model",
        orientation="h",
        color_discrete_sequence=["#1e40af"],
        labels={"Total": "판매대수", "Model": ""},
    )
    fig_bar_domestic.update_layout(height=380, showlegend=False, xaxis_tickformat=",")
    st.plotly_chart(fig_bar_domestic, use_container_width=True)

with col_top2:
    st.markdown("**해외 공장 - Top 10 모델**")
    top_export = export.nlargest(10, "Total")[["Model", "Total"]].sort_values("Total")
    fig_bar_export = px.bar(
        top_export,
        x="Total",
        y="Model",
        orientation="h",
        color_discrete_sequence=["#f97316"],
        labels={"Total": "판매대수", "Model": ""},
    )
    fig_bar_export.update_layout(height=380, showlegend=False, xaxis_tickformat=",")
    st.plotly_chart(fig_bar_export, use_container_width=True)

st.markdown("---")

# ==================== 6. 모델별 월별 성과 분석 ====================
st.markdown('<h2 class="section-title">모델별 월별 성과 분석</h2>', unsafe_allow_html=True)

selected_model = st.selectbox(
    "분석할 모델 선택:", options=sorted(df["Model"].unique()), index=0
)

model_data = df[df["Model"] == selected_model]

if len(model_data) > 0:
    col_model1, col_model2 = st.columns(2, gap="large")

    with col_model1:
        st.markdown(f"**{selected_model} - 월별 판매량 (누적 막대 그래프)**")

        model_monthly = model_data.groupby("Plant")[months].sum()

        fig_stacked = go.Figure()
        for plant in model_monthly.index:
            fig_stacked.add_trace(
                go.Bar(
                    x=months_ko,
                    y=model_monthly.loc[plant].values,
                    name=plant,
                    hovertemplate="<b>"
                    + plant
                    + "</b><br>%{x}: %{y:,.0f}대<extra></extra>",
                )
            )

        fig_stacked.update_layout(
            barmode="stack",
            height=380,
            yaxis_title="판매대수 (대)",
            yaxis_tickformat=",",
        )
        st.plotly_chart(fig_stacked, use_container_width=True)

    with col_model2:
        st.markdown(f"**{selected_model} - 공장별 월별 판매 상세**")
        model_summary = model_data[["Plant"] + months].set_index("Plant")[months].T
        model_summary.index = months_ko
        st.dataframe(model_summary.astype(int), use_container_width=True)

st.markdown("---")

# ==================== 7. 상세 데이터 테이블 ====================
if layout_mode:
    st.markdown('<h2 class="section-title">상세 데이터</h2>', unsafe_allow_html=True)

    display_cols = ["Plant", "Factory", "Model"] + months + ["Total"]
    display_df = df[display_cols].copy()
    display_df = display_df.sort_values(["Plant", "Total"], ascending=[True, False])

    st.dataframe(display_df, use_container_width=True, hide_index=True)

    csv = df.to_csv(index=False)
    st.download_button(
        label="다운로드 (CSV)",
        data=csv,
        file_name=f"Global_Sales_{selected_year}.csv",
        mime="text/csv",
    )


# ==================== 8. Gemini AI 내년도 수요 예측 ====================
from google import genai
from google.genai.types import HttpOptions

st.markdown('<h2 class="section-title">AI 수요 예측 분석 (Gemini 2.5)</h2>', unsafe_allow_html=True)

st.markdown("""
<style>
    .gemini-card {
        background: linear-gradient(135deg, #ecfdf5 0%, #d1fae5 100%);
        padding: 20px 24px;
        border-radius: 12px;
        border-left: 5px solid #10b981;
        margin-bottom: 20px;
        box-shadow: 0 4px 14px rgba(16, 185, 129, 0.2);
    }
    .gemini-title {
        font-size: 18px;
        font-weight: 700;
        color: #065f46;
        margin-bottom: 12px;
    }
    .gemini-loading {
        background: linear-gradient(90deg, #10b981 0%, #34d399 50%, #10b981 100%);
        background-size: 200% 100%;
        animation: loading 1.5s infinite;
    }
    @keyframes loading {
        0% { background-position: 200% 0; }
        100% { background-position: -200% 0; }
    }
</style>
""", unsafe_allow_html=True)

# 버튼
if st.button("Gemini 2.5로 내년도 수요 예측 분석", type="primary", use_container_width=True):
    st.session_state.show_gemini = True
    st.session_state.gemini_running = True
    st.session_state.pop("gemini_analysis", None)
    st.session_state.pop("gemini_summary", None)
    st.session_state.pop("gemini_error", None)
    st.session_state.gemini_stream_text = ""
    st.rerun()
else:
    st.session_state.show_gemini = st.session_state.get("show_gemini", False)
    st.session_state.gemini_running = st.session_state.get("gemini_running", False)

if st.session_state.show_gemini:

    # 공통 데이터 요약
    top_domestic_model = domestic.nlargest(1, "Total")["Model"].iloc[0]
    top_export_model = export.nlargest(1, "Total")["Model"].iloc[0]

    monthly_domestic = domestic[months].sum()
    monthly_export = export[months].sum()

    summary_data = f"""
현대차 {selected_year}년 글로벌 판매 데이터 요약

[전체 실적]
- 총 판매량: {grand_total:,.0f}대
- 국내 공장: {domestic_total:,.0f}대 ({domestic_total/grand_total*100:.1f}%)
- 해외 공장: {export_total:,.0f}대 ({export_total/grand_total*100:.1f}%)
- 등록 모델 수: {df["Model"].nunique()}종

[국내 Top 모델]
- 1위: {top_domestic_model} ({domestic.nlargest(1, "Total")["Total"].iloc[0]:,.0f}대)

[해외 Top 모델]
- 1위: {top_export_model} ({export.nlargest(1, "Total")["Total"].iloc[0]:,.0f}대)

[국내 월별 추이]
- 1월: {monthly_domestic.iloc[0]:,.0f}대
- 12월: {monthly_domestic.iloc[-1]:,.0f}대
- 성장률: {((monthly_domestic.iloc[-1] - monthly_domestic.iloc[0]) / max(monthly_domestic.iloc[0],1) * 100):+.1f}%
"""

    # 좀 더 짧게 다이어트한 실제 프롬프트
    prompt = f"""
현대차 {selected_year}년 글로벌 판매 요약:
- 총 판매: {grand_total:,.0f}대
- 국내/해외 비중: {domestic_total/grand_total*100:.1f}% / {export_total/grand_total*100:.1f}%
- 국내 1위: {top_domestic_model}
- 해외 1위: {top_export_model}

위 데이터를 기반으로 {selected_year+1}년 수요를 예측하고,
1) 국내·해외 예상 판매량과 성장률,
2) 상위 3개 모델 역할과 리스크,
3) 주요 리스크와 대응전략,
4) 경영진이 참고할 핵심 인사이트 3가지
를 한국어로 요약해 주세요.
"""

    # 1) Gemini 스트리밍 호출
    if st.session_state.gemini_running:
        st.markdown(f"""
        <div class="gemini-card">
            <div class="gemini-title">Gemini 2.5 분석 중</div>
            <div style="font-size: 14px; color: #065f46;">
                최신 데이터를 바탕으로 {selected_year+1}년 수요를 분석하고 있습니다.
            </div>
            <div class="gemini-loading" style="height: 4px; border-radius: 2px; margin-top: 12px;"></div>
        </div>
        """, unsafe_allow_html=True)

        placeholder = st.empty()

        try:
            if GEMINI_API_KEY and len(GEMINI_API_KEY) > 30:
                client = genai.Client(
                    api_key=GEMINI_API_KEY,
                    http_options=HttpOptions(timeout=60_000)
                )

                st.session_state.gemini_stream_text = ""

                stream = client.models.generate_content_stream(
                    model="gemini-2.5-flash",
                    contents=prompt,
                )

                for chunk in stream:
                    part = getattr(chunk, "text", "") or ""
                    st.session_state.gemini_stream_text += part
                    placeholder.markdown(st.session_state.gemini_stream_text)

                ai_text = st.session_state.gemini_stream_text.strip()

                st.session_state.gemini_running = False
                st.session_state.gemini_analysis = ai_text
                st.session_state.gemini_summary = summary_data
                st.session_state.gemini_error = None
                st.rerun()
            else:
                st.session_state.gemini_running = False
                st.session_state.gemini_error = "API_KEY_MISSING"
                st.session_state.gemini_summary = summary_data
                st.rerun()

        except Exception as e:
            st.session_state.gemini_running = False
            st.session_state.gemini_error = str(e)
            st.session_state.gemini_summary = summary_data
            st.rerun()

    # 2) Gemini 결과 + 내부 예측 + 다운로드
    else:
        err = st.session_state.get("gemini_error")
        ai_text = st.session_state.get("gemini_analysis")

        # 내부 예측 (항상 계산)
        growth_rate = 0.08
        current_domestic = float(domestic_total)
        current_export = float(export_total)
        current_total = float(grand_total)

        forecast_domestic = int(current_domestic * (1 + growth_rate))
        forecast_export = int(current_export * (1 + growth_rate))
        forecast_total = forecast_domestic + forecast_export
        total_growth_pct = (forecast_total - current_total) / max(current_total, 1) * 100

        forecast_dict = {
            'current_domestic': current_domestic,
            'current_export': current_export,
            'current_total': current_total,
            'forecast_domestic': forecast_domestic,
            'forecast_export': forecast_export,
            'forecast_total': forecast_total,
            'total_growth_pct': total_growth_pct
        }

        # 상태 카드
        if ai_text:
            st.markdown(f"""
            <div class="gemini-card">
                <div class="gemini-title">Gemini 2.5 분석 완료</div>
                <div style="font-size: 14px; color: #065f46;">
                    {selected_year+1}년 현대차 글로벌 수요 예측 결과입니다.
                </div>
            </div>
            """, unsafe_allow_html=True)

            st.markdown("#### 분석 기준 데이터 (요약)")
            st.markdown(f"> {st.session_state.gemini_summary.replace('\\n', '\\n> ')}")

            st.markdown("#### Gemini 2.5의 내년도 수요 예측")
            with st.expander("Gemini 2.5 전체 분석 보기", expanded=True):
                st.markdown(ai_text)
        else:
            st.markdown(f"""
            <div class="gemini-card">
                <div class="gemini-title">Gemini 분석 대신 내부 예측을 사용합니다</div>
                <div style="font-size: 13px; color: #b91c1c; margin-top: 4px;">
                    외부 Gemini API 호출에 실패하여, 대시보드 데이터 기반으로 자체 예측을 계산합니다.
                </div>
            </div>
            """, unsafe_allow_html=True)

            if err:
                if "429" in err or "quota" in err.lower():
                    st.info("현재 Google Gemini 프로젝트의 무료/할당량이 0으로 설정되어 있어 외부 모델을 호출할 수 없는 상태입니다. 콘솔에서 요금제·쿼터를 조정해야 계속 사용할 수 있습니다.")
                elif "404" in err:
                    st.info("지정한 Gemini 모델이 이 API 버전 또는 프로젝트에서 허용되지 않습니다. Google AI Studio에서 사용 가능한 모델 목록을 확인해야 합니다.")
                else:
                    st.text_area("Gemini 에러 상세", err, height=100)

        # KPI
        col_f1, col_f2, col_f3 = st.columns(3)
        with col_f1:
            st.metric("국내 공장 내년도 예상", f"{forecast_domestic:,}대", f"+{forecast_domestic-current_domestic:,.0f}대")
        with col_f2:
            st.metric("해외 공장 내년도 예상", f"{forecast_export:,}대", f"+{forecast_export-current_export:,.0f}대")
        with col_f3:
            st.metric("총 내년도 예상", f"{forecast_total:,}대", f"{total_growth_pct:+.1f}%")

        # 월별 예측 그래프
        st.markdown("#### 월별 수요 예측 (내부 계산)")
        dom_ratio = monthly_domestic.values / max(current_domestic, 1)
        exp_ratio = monthly_export.values / max(current_export, 1)

        dom_month_forecast = forecast_domestic * dom_ratio
        exp_month_forecast = forecast_export * exp_ratio

        fig_fc = go.Figure()
        fig_fc.add_trace(go.Scatter(
            x=months_ko, y=monthly_domestic.values,
            name=f"국내 {selected_year}년 (실적)",
            mode="lines+markers",
            line=dict(color="#1d4ed8", width=3),
            marker=dict(size=6),
        ))
        fig_fc.add_trace(go.Scatter(
            x=months_ko, y=dom_month_forecast,
            name=f"국내 {selected_year+1}년 (예상)",
            mode="lines+markers",
            line=dict(color="#1d4ed8", width=2, dash="dash"),
            marker=dict(size=6, symbol="diamond-open"),
        ))
        fig_fc.add_trace(go.Scatter(
            x=months_ko, y=monthly_export.values,
            name=f"해외 {selected_year}년 (실적)",
            mode="lines+markers",
            line=dict(color="#f97316", width=3),
            marker=dict(size=6),
        ))
        fig_fc.add_trace(go.Scatter(
            x=months_ko, y=exp_month_forecast,
            name=f"해외 {selected_year+1}년 (예상)",
            mode="lines+markers",
            line=dict(color="#f97316", width=2, dash="dash"),
            marker=dict(size=6, symbol="diamond-open"),
        ))

        fig_fc.update_layout(
            height=380,
            hovermode="x unified",
            yaxis_title="판매대수 (대)",
            yaxis_tickformat=",",
            legend=dict(x=0.01, y=0.99, bgcolor="rgba(255,255,255,0.9)")
        )
        st.plotly_chart(fig_fc, use_container_width=True)

        # PDF / Excel 다운로드
        st.markdown("---")
        st.markdown("#### 분석 결과 다운로드")

        col_dl1, col_dl2 = st.columns(2, gap="medium")

        with col_dl1:
            try:
                from reportlab.lib.pagesizes import A4
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
                from reportlab.lib.units import inch
                from reportlab.lib import colors
                from io import BytesIO
                from datetime import datetime

                def create_forecast_pdf(forecast_data, summary_data, ai_analysis=None):
                    buffer = BytesIO()
                    doc = SimpleDocTemplate(
                        buffer,
                        pagesize=A4,
                        rightMargin=0.75*inch,
                        leftMargin=0.75*inch,
                        topMargin=0.75*inch,
                        bottomMargin=0.75*inch
                    )

                    styles = getSampleStyleSheet()
                    title_style = ParagraphStyle(
                        'CustomTitle',
                        parent=styles['Heading1'],
                        fontSize=18,
                        textColor=colors.HexColor('#111827'),
                        spaceAfter=12,
                        fontName='Helvetica-Bold'
                    )
                    heading_style = ParagraphStyle(
                        'CustomHeading',
                        parent=styles['Heading2'],
                        fontSize=13,
                        textColor=colors.HexColor('#1e40af'),
                        spaceAfter=8,
                        spaceBefore=8,
                        fontName='Helvetica-Bold'
                    )
                    body_style = ParagraphStyle(
                        'CustomBody',
                        parent=styles['BodyText'],
                        fontSize=10,
                        spaceAfter=6
                    )

                    elements = []
                    elements.append(Paragraph(
                        f"현대차 {selected_year+1}년 수요 예측 분석 보고서",
                        title_style
                    ))
                    elements.append(Paragraph(
                        f"작성일: {datetime.now().strftime('%Y년 %m월 %d일')}",
                        body_style
                    ))
                    elements.append(Spacer(1, 0.3*inch))

                    elements.append(Paragraph("분석 기준 데이터", heading_style))
                    summary_lines = summary_data.strip().split('\n')
                    for line in summary_lines:
                        if line.strip():
                            elements.append(Paragraph(line, body_style))
                    elements.append(Spacer(1, 0.2*inch))

                    fd = forecast_data
                    forecast_table_data = [
                        ['구분', '올해 실적', '내년도 예상', '증감', '성장률'],
                        ['국내 공장',
                         f"{int(fd['current_domestic']):,}대",
                         f"{fd['forecast_domestic']:,}대",
                         f"+{fd['forecast_domestic'] - int(fd['current_domestic']):,}대",
                         f"{(fd['forecast_domestic'] - int(fd['current_domestic'])) / max(int(fd['current_domestic']), 1) * 100:+.1f}%"],
                        ['해외 공장',
                         f"{int(fd['current_export']):,}대",
                         f"{fd['forecast_export']:,}대",
                         f"+{fd['forecast_export'] - int(fd['current_export']):,}대",
                         f"{(fd['forecast_export'] - int(fd['current_export'])) / max(int(fd['current_export']), 1) * 100:+.1f}%"],
                        ['합계',
                         f"{int(fd['current_total']):,}대",
                         f"{fd['forecast_total']:,}대",
                         f"+{fd['forecast_total'] - int(fd['current_total']):,}대",
                         f"{fd['total_growth_pct']:+.1f}%"]
                    ]

                    forecast_table = Table(
                        forecast_table_data,
                        colWidths=[1.2*inch, 1.2*inch, 1.2*inch, 1*inch, 0.8*inch]
                    )
                    forecast_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 10),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e0f2fe')),
                        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black)
                    ]))
                    elements.append(forecast_table)
                    elements.append(Spacer(1, 0.3*inch))

                    if ai_analysis:
                        elements.append(PageBreak())
                        elements.append(Paragraph("Gemini 2.5 상세 분석", heading_style))
                        analysis_lines = ai_analysis.split('\n')
                        for line in analysis_lines[:50]:
                            if line.strip():
                                elements.append(Paragraph(line, body_style))

                    elements.append(Spacer(1, 0.5*inch))
                    elements.append(Paragraph(
                        "본 보고서는 현재 판매 데이터를 기반으로 작성된 예측 분석입니다.",
                        body_style
                    ))

                    doc.build(elements)
                    buffer.seek(0)
                    return buffer

                pdf_buffer = create_forecast_pdf(
                    forecast_dict,
                    st.session_state.get("gemini_summary", summary_data),
                    ai_text if ai_text else None
                )

                st.download_button(
                    label="PDF 다운로드 (분석 보고서)",
                    data=pdf_buffer,
                    file_name=f"현대차_{selected_year+1}년_수요예측_{datetime.now().strftime('%Y%m%d')}.pdf",
                    mime="application/pdf",
                    use_container_width=True
                )

            except ImportError:
                st.warning("PDF 생성을 위해 reportlab 패키지가 필요합니다.")
                st.code("pip install reportlab")

        with col_dl2:
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
                from io import BytesIO
                from datetime import datetime as _dt3

                def create_forecast_excel(forecast_data, monthly_data):
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "수요예측"

                    headers = ['구분', '올해 실적', '내년도 예상', '증감', '성장률']
                    ws.append(headers)

                    header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF")

                    for col_num, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col_num)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                    fd = forecast_data
                    data_rows = [
                        ['국내 공장',
                         f"{int(fd['current_domestic']):,}",
                         f"{fd['forecast_domestic']:,}",
                         f"+{fd['forecast_domestic'] - int(fd['current_domestic']):,}",
                         f"{(fd['forecast_domestic'] - int(fd['current_domestic'])) / max(int(fd['current_domestic']), 1) * 100:+.1f}%"],
                        ['해외 공장',
                         f"{int(fd['current_export']):,}",
                         f"{fd['forecast_export']:,}",
                         f"+{fd['forecast_export'] - int(fd['current_export']):,}",
                         f"{(fd['forecast_export'] - int(fd['current_export'])) / max(int(fd['current_export']), 1) * 100:+.1f}%"],
                        ['합계',
                         f"{int(fd['current_total']):,}",
                         f"{fd['forecast_total']:,}",
                         f"+{fd['forecast_total'] - int(fd['current_total']):,}",
                         f"{fd['total_growth_pct']:+.1f}%"]
                    ]

                    for row_data in data_rows:
                        ws.append(row_data)

                    ws.column_dimensions['A'].width = 15
                    for col in ['B', 'C', 'D', 'E']:
                        ws.column_dimensions[col].width = 18

                    ws2 = wb.create_sheet("월별예측")
                    ws2.append(['월', f'{selected_year}년 국내', f'{selected_year+1}년 국내',
                                f'{selected_year}년 해외', f'{selected_year+1}년 해외', '합계'])

                    for idx, month in enumerate(months_ko):
                        dom_current = int(monthly_data['dom_current'][idx])
                        dom_forecast = int(monthly_data['dom_forecast'][idx])
                        exp_current = int(monthly_data['exp_current'][idx])
                        exp_forecast = int(monthly_data['exp_forecast'][idx])

                        ws2.append([
                            month, dom_current, dom_forecast,
                            exp_current, exp_forecast,
                            dom_forecast + exp_forecast
                        ])

                    for row in ws2.iter_rows(min_row=1, max_row=13, min_col=1, max_col=6):
                        for cell in row:
                            cell.alignment = Alignment(horizontal='center')
                            if cell.row == 1:
                                cell.fill = header_fill
                                cell.font = header_font

                    ws2.column_dimensions['A'].width = 12
                    for col in ['B', 'C', 'D', 'E', 'F']:
                        ws2.column_dimensions[col].width = 16

                    return wb

                monthly_dict = {
                    'dom_current': monthly_domestic.values,
                    'dom_forecast': dom_month_forecast,
                    'exp_current': monthly_export.values,
                    'exp_forecast': exp_month_forecast
                }

                excel_wb = create_forecast_excel(forecast_dict, monthly_dict)

                excel_buffer = BytesIO()
                excel_wb.save(excel_buffer)
                excel_buffer.seek(0)

                st.download_button(
                    label="Excel 다운로드 (예측 데이터)",
                    data=excel_buffer,
                    file_name=f"현대차_{selected_year+1}년_수요예측_{_dt3.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

            except ImportError:
                st.warning("Excel 생성을 위해 openpyxl 패키지가 필요합니다.")
                st.code("pip install openpyxl")

        # 공통: 닫기 버튼
        if st.button("예측 분석 닫기", type="secondary", use_container_width=True):
            for k in ["show_gemini", "gemini_running", "gemini_analysis", "gemini_summary", "gemini_error", "gemini_stream_text"]:
                st.session_state.pop(k, None)
            st.rerun()
